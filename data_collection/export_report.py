import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import os
import csv

from datetime import datetime

IMAGE_DIR  = "test_instellingen"
CSV_PATH   = os.path.join(IMAGE_DIR, "resultaten.csv")
OUTPUT_DIR = f"rapport_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
os.makedirs(OUTPUT_DIR, exist_ok=True)
print(f"Rapport map: {OUTPUT_DIR}")

# --- CSV inlezen (gegenereerd door test_auto.py) ---
if not os.path.exists(CSV_PATH):
    print(f"CSV niet gevonden: {CSV_PATH}")
    print("Voer eerst test_auto.py uit om de resultaten te genereren.")
    exit()

results = []
with open(CSV_PATH, encoding='utf-8') as f:
    for row in csv.DictReader(f):
        results.append({
            'filename':      row['filename'],
            'exposure':      int(row['exposure']),
            'gain':          int(row['gain']),
            'sharp_param':   int(row['sharpness']),
            'contrast_param':int(row['contrast']),
            'combined':      float(row['combined_score']),
            'defect':        float(row.get('defect_score', row['sharpness_score'])),
            'sharpness':     float(row['sharpness_score']),
            'brightness':    float(row['brightness']),
            'noise':         float(row['noise']),
            'over':          float(row['overexposed']),
        })

results.sort(key=lambda r: r['combined'], reverse=True)
print(f"{len(results)} resultaten ingelezen uit {CSV_PATH}")

# --- Excel export ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Resultaten"

headers = ["#", "Bestand", "Exposure", "Gain", "Sharpness", "Contrast",
           "Combined Score", "Defect Score", "Sharpness Score", "Brightness", "Noise", "Overexposed %"]
ws.append(headers)

header_fill = PatternFill("solid", fgColor="2E75B6")
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center")

green = PatternFill("solid", fgColor="C6EFCE")
red   = PatternFill("solid", fgColor="FFC7CE")
gray  = PatternFill("solid", fgColor="F2F2F2")

for rank, r in enumerate(results, 1):
    ws.append([rank, r['filename'], r['exposure'], r['gain'],
               r['sharp_param'], r['contrast_param'],
               r['combined'], r['defect'], r['sharpness'], r['brightness'], r['noise'], r['over']])
    row_idx = rank + 1
    fill = green if rank == 1 else red if rank == len(results) else (gray if rank % 2 == 0 else None)
    if fill:
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).fill = fill
    if r['brightness'] < 60 or r['brightness'] > 200:
        ws.cell(row=row_idx, column=10).font = Font(color="FF0000", bold=True)

for i, width in enumerate([5, 35, 10, 6, 11, 10, 15, 14, 16, 12, 8, 14], 1):
    ws.column_dimensions[get_column_letter(i)].width = width
ws.freeze_panes = "A2"

excel_path = os.path.join(OUTPUT_DIR, "resultaten.xlsx")
wb.save(excel_path)
print(f"Excel opgeslagen: {excel_path}")

# --- Grafieken als PNG ---
fig, axes = plt.subplots(2, 2, figsize=(14, 10))
fig.suptitle("Gemiddelde defect score per parameterwaarde\n(hogere score = barst/defect beter zichtbaar)", fontsize=12)

for (key, label), ax in zip([('exposure','Exposure'), ('gain','Gain'),
                               ('sharp_param','Sharpness'), ('contrast_param','Contrast')],
                              axes.flatten()):
    groups = {}
    for r in results:
        groups.setdefault(r[key], []).append(r['defect'])
    vals  = sorted(groups)
    means = [np.mean(groups[v]) for v in vals]
    ax.bar([str(v) for v in vals], means, color='steelblue')
    ax.set_title(label)
    ax.set_ylabel("Gem. defect score")
    ax.grid(axis='y', alpha=0.4)
    for v, mean in zip(vals, means):
        ax.text(str(v), mean + 1, f"{mean:.0f}", ha='center', fontsize=9)

fig.tight_layout()
graph_path = os.path.join(OUTPUT_DIR, "grafiek_parameters.png")
fig.savefig(graph_path, dpi=150, bbox_inches='tight')
plt.close(fig)
print(f"Grafiek opgeslagen: {graph_path}")

# --- Aanbeveling ---
best = results[0]
print("\n" + "=" * 50)
print("  BESTE INSTELLINGEN")
print("=" * 50)
print(f"  Exposure  = {best['exposure']}")
print(f"  Gain      = {best['gain']}")
print(f"  Sharpness = {best['sharp_param']}")
print(f"  Contrast  = {best['contrast_param']}")
print(f"  Combined score = {best['combined']:.1f}")
print(f"  Defect score   = {best['defect']:.1f}  (zichtbaarheid barst/defect)")
print(f"  Brightness     = {best['brightness']:.1f}  (ideaal: 80-180)")
print(f"  Noise          = {best['noise']:.2f}  (< 8 is goed)")
print(f"  Overexp.       = {best['over']:.2f}%  (< 1% is goed)")
